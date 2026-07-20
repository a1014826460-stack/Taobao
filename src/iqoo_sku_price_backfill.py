"""Fill an iQOO SKU workbook from locally captured Tmall detail records."""

from __future__ import annotations

import argparse
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
import json
import re
import sqlite3
from pathlib import Path
from typing import Any
from urllib.parse import urlsplit

import openpyxl


@dataclass(frozen=True)
class SkuMatch:
    item_id: str
    sku_id: str
    price: str
    installment_periods: int | None = None


@dataclass(frozen=True)
class SupplementResult:
    attempted_ids: list[str]
    paused: bool


def normalize_text(value: object) -> str:
    return "".join(str(value or "").lower().split())


def _sku_property_names(sku_base: dict[str, Any], prop_path: str) -> list[str]:
    values_by_path: dict[str, str] = {}
    for prop in sku_base.get("props") or []:
        if not isinstance(prop, dict):
            continue
        pid = str(prop.get("pid") or "")
        for value in prop.get("values") or []:
            if isinstance(value, dict):
                values_by_path[f"{pid}:{value.get('vid')}"] = str(value.get("name") or "")
    return [values_by_path.get(part, "") for part in str(prop_path).split(";")]


def normalize_sku_value(value: object) -> str:
    normalized = normalize_text(value)
    normalized = normalized.replace("gb", "g").replace("tb", "t")
    normalized = re.sub(r"(?<=\d)g(?=\+|$)", "", normalized)
    return normalized


def match_sku(
    detail: dict[str, Any], model: str, configuration: str, color: str
) -> SkuMatch | None:
    """Return a SKU only if the three workbook fields match it exactly."""
    if normalize_text(model) not in normalize_text(detail.get("title")):
        return None
    sku_base = detail.get("sku_base") or {}
    sku_core = detail.get("sku_core") or {}
    sku_info = sku_core.get("sku2info") or {}
    expected = {normalize_sku_value(configuration), normalize_sku_value(color)}
    for sku in sku_base.get("skus") or []:
        if not isinstance(sku, dict):
            continue
        names = {
            normalize_sku_value(name)
            for name in _sku_property_names(sku_base, sku.get("propPath", ""))
        }
        if not expected.issubset(names):
            continue
        sku_id = str(sku.get("skuId") or "")
        info = sku_info.get(sku_id) or {}
        price_data = info.get("price") if isinstance(info, dict) else {}
        price = str((price_data or {}).get("priceText") or "")
        if price:
            terms = extract_installment_periods(info)
            if terms is None:
                terms = detail.get("installment_periods")
            return SkuMatch(str(detail.get("item_id") or ""), sku_id, price, terms)
    return None


def extract_installment_periods(value: Any) -> int | None:
    if isinstance(value, dict):
        for key, child in value.items():
            if key.lower() in {"installmentperiods", "installmentnum", "period", "periods"}:
                try:
                    periods = int(str(child))
                    if periods > 0:
                        return periods
                except (TypeError, ValueError):
                    pass
            periods = extract_installment_periods(child)
            if periods:
                return periods
    elif isinstance(value, list):
        for child in value:
            periods = extract_installment_periods(child)
            if periods:
                return periods
    return None


def extract_page_installment_periods(value: Any) -> int | None:
    """Read an explicit X期免息 term from captured page text."""
    text = json.dumps(value, ensure_ascii=False)
    terms = [int(match) for match in re.findall(r"(\d+)\s*期免息", text)]
    return max(terms) if terms else None


def extract_page_installment_text(value: Any) -> str:
    text = json.dumps(value, ensure_ascii=False)
    match = re.search(r"(\d+)\s*期免息\s*约\s*¥\s*([\d.]+)\s*/\s*期", text)
    return f"¥{match.group(2)} × {match.group(1)}期" if match else ""


def normalize_installment_text(value: object) -> str:
    match = re.search(r"(\d+)\s*期免息\s*约\s*¥\s*([\d.]+)\s*/\s*期", str(value or ""))
    return f"¥{match.group(2)} × {match.group(1)}期" if match else ""


def extract_coupon_price(value: dict[str, Any]) -> str:
    price_vo = ((value.get("componentsVO") or {}).get("priceVO") or {})
    return str(((price_vo.get("extraPrice") or {}).get("priceText")) or "")


def _clean_price_text(value: object) -> str:
    match = re.search(r"\d+(?:\.\d+)?", str(value or ""))
    return match.group(0) if match else ""


def _sku_price(info: dict[str, Any], fallback_price: object) -> tuple[str, int]:
    """Return the most specific captured price for one SKU.

    `extraPrice` is a product-level starting price; `subPrice` is the
    selection-specific discounted price returned after choosing a SKU.
    """
    sub_price = _clean_price_text((info.get("subPrice") or {}).get("priceText"))
    if sub_price:
        return sub_price, 3
    price = _clean_price_text((info.get("price") or {}).get("priceText"))
    if price:
        return price, 2
    return _clean_price_text(fallback_price), 1


def format_installment(price: str, periods: int | None) -> str:
    if not periods or periods <= 0:
        return ""
    try:
        per_period = (Decimal(str(price)) / Decimal(periods)).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )
    except (InvalidOperation, ValueError):
        return ""
    return f"¥{per_period} × {periods}期"


def load_shop_candidates(db_path: str | Path, shop_url: str, model: str) -> list[str]:
    connection = sqlite3.connect(db_path)
    try:
        rows = connection.execute(
            "SELECT shop_url, item_id, title FROM tmall_shop_items ORDER BY item_id"
        ).fetchall()
        has_page_table = connection.execute(
            "SELECT 1 FROM sqlite_master WHERE type = 'table' AND name = 'tmall_shop_pages'"
        ).fetchone()
        page_rows = (
            connection.execute("SELECT shop_url, raw_json FROM tmall_shop_pages").fetchall()
            if has_page_table
            else []
        )
    finally:
        connection.close()
    requested = urlsplit(shop_url)
    requested_identity = (requested.scheme, requested.netloc, requested.path)
    expected = normalize_text(model)
    model_tokens = [token for token in re.split(r"\s+", str(model).strip()) if token]
    pattern = re.compile(
        r"(?<![a-z0-9])" + r"\s*".join(re.escape(token) for token in model_tokens) + r"(?![a-z0-9])",
        re.IGNORECASE,
    )

    def title_matches(title: object) -> bool:
        return bool(pattern.search(str(title or "")))

    candidate_ids = [
        str(item_id)
        for saved_url, item_id, title in rows
        if (urlsplit(saved_url).scheme, urlsplit(saved_url).netloc, urlsplit(saved_url).path)
        == requested_identity
        and title_matches(title)
    ]
    if candidate_ids:
        return candidate_ids
    for saved_url, raw_json in page_rows:
        saved = urlsplit(saved_url)
        if (saved.scheme, saved.netloc, saved.path) != requested_identity:
            continue
        try:
            html = json.loads(raw_json)
        except (TypeError, ValueError):
            continue
        if not isinstance(html, str):
            continue
        for match in re.finditer(
            r'<dl\b[^>]*\bdata-id\s*=\s*(["\'])(?P<item_id>[^"\']+)\1[^>]*>(?P<body>.*?)</dl>',
            html,
            re.IGNORECASE | re.DOTALL,
        ):
            item_text = re.sub(r"<[^>]+>", " ", match.group("body"))
            if title_matches(item_text):
                candidate_ids.append(match.group("item_id"))
    return list(dict.fromkeys(candidate_ids))


def load_detail_records(db_path: str | Path) -> list[dict[str, Any]]:
    connection = sqlite3.connect(db_path)
    connection.row_factory = sqlite3.Row
    try:
        rows = connection.execute(
            "SELECT item_id, title, loader_data_json FROM taobao_item_details WHERE status = 'success'"
        ).fetchall()
    finally:
        connection.close()
    records = []
    for row in rows:
        try:
            payload = json.loads(row["loader_data_json"])
            result = payload.get("home", {}).get("data", {}).get("res", {})
            records.append(
                {
                    "item_id": str(row["item_id"]),
                    "title": row["title"] or result.get("item", {}).get("title", ""),
                    "sku_base": result.get("skuBase") or {},
                    "sku_core": result.get("skuCore") or {},
                    "installment_periods": extract_page_installment_periods(result),
                    "installment_text": extract_page_installment_text(result),
                    "coupon_price": extract_coupon_price(result),
                }
            )
        except (TypeError, ValueError, AttributeError):
            continue
    return records


REPORT_HEADERS = [
    "Excel行号",
    "机型",
    "配置",
    "颜色",
    "候选商品ID",
    "价格状态",
    "免息状态",
    "原因",
]

PHONE_EXPORT_HEADERS = [
    "店铺",
    "网页版链接",
    "品牌(*)",
    "机型(*)",
    "配置(*)",
    "颜色(*)",
    "价格",
    "免息分期",
]


def phone_model_from_title(title: str) -> str | None:
    match = re.search(
        r"\biQOO\s+(Z\d+(?:\s*Turbo\+?|[xi])?|Neo\d+|\d+\s*Ultra|\d+T?)(?![a-z0-9])",
        title,
        re.IGNORECASE,
    )
    if not match:
        return None
    suffix = re.sub(r"\s+", " ", match.group(1)).strip()
    return "iQOO " + suffix


def _phone_sku_values(detail: dict[str, Any], sku: dict[str, Any]) -> tuple[str, str]:
    sku_base = detail.get("sku_base") or {}
    values_by_path: dict[str, tuple[str, str]] = {}
    for prop in sku_base.get("props") or []:
        if not isinstance(prop, dict):
            continue
        pid = str(prop.get("pid") or "")
        prop_name = str(prop.get("name") or "")
        for value in prop.get("values") or []:
            if isinstance(value, dict):
                values_by_path[f"{pid}:{value.get('vid')}"] = (
                    prop_name,
                    str(value.get("name") or ""),
                )
    color = configuration = ""
    for part in str(sku.get("propPath") or "").split(";"):
        prop_name, value_name = values_by_path.get(part, ("", ""))
        if "颜色" in prop_name:
            color = value_name
        elif any(token in prop_name for token in ("容量", "内存", "版本", "存储")):
            configuration = value_name
    return configuration, color


def build_phone_export_rows(details: list[dict[str, Any]], shop_url: str) -> list[dict[str, str]]:
    rows_by_identity: dict[tuple[str, str, str, str], tuple[int, dict[str, str]]] = {}
    for detail in details:
        model = phone_model_from_title(str(detail.get("title") or ""))
        if not model:
            continue
        sku_base = detail.get("sku_base") or {}
        sku_info = (detail.get("sku_core") or {}).get("sku2info") or {}
        for sku in sku_base.get("skus") or []:
            if not isinstance(sku, dict):
                continue
            configuration, color = _phone_sku_values(detail, sku)
            if not configuration or not color:
                continue
            info = sku_info.get(str(sku.get("skuId") or "")) or {}
            price, price_specificity = _sku_price(info, detail.get("coupon_price"))
            if not price:
                continue
            normalized_configuration = re.sub(
                r"^(\d+)GB\+", r"\1+", normalize_text(configuration).upper()
            ).replace("GB", "G").replace("TB", "T")
            identity = ("iQOO", model, normalized_configuration, color)
            previous = rows_by_identity.get(identity)
            if previous and previous[0] >= price_specificity:
                continue
            installment = format_installment(price, detail.get("installment_periods"))
            if price_specificity == 1:
                installment = normalize_installment_text(detail.get("installment_text")) or installment
            rows_by_identity[identity] = (
                price_specificity,
                {
                    "店铺": "iQOO手机官方旗舰店",
                    "网页版链接": shop_url,
                    "品牌(*)": identity[0],
                    "机型(*)": identity[1],
                    "配置(*)": identity[2],
                    "颜色(*)": identity[3],
                    "价格": price,
                    "免息分期": installment,
                },
            )
    return [entry[1] for entry in rows_by_identity.values()]


def write_phone_export(output_path: str | Path, rows: list[dict[str, str]]) -> int:
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "iQOO手机SKU"
    sheet.append(PHONE_EXPORT_HEADERS)
    for row in rows:
        sheet.append([row.get(header, "") for header in PHONE_EXPORT_HEADERS])
    sheet.freeze_panes = "A2"
    for column in sheet.columns:
        letter = column[0].column_letter
        sheet.column_dimensions[letter].width = min(
            max(12, max(len(str(cell.value or "")) for cell in column[:100]) + 2), 55
        )
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    book.save(output_path)
    return len(rows)


def phone_detail_coverage(details: list[dict[str, Any]]) -> tuple[int, int, set[str]]:
    """Return phone product count, price-ready SKU count, and seen phone IDs."""
    phone_ids: set[str] = set()
    price_ready_skus = 0
    for detail in details:
        if not phone_model_from_title(str(detail.get("title") or "")):
            continue
        phone_ids.add(str(detail.get("item_id") or ""))
        sku_info = (detail.get("sku_core") or {}).get("sku2info") or {}
        for sku in (detail.get("sku_base") or {}).get("skus") or []:
            if not isinstance(sku, dict):
                continue
            configuration, color = _phone_sku_values(detail, sku)
            info = sku_info.get(str(sku.get("skuId") or "")) or {}
            if configuration and color and (info.get("price") or {}).get("priceText"):
                price_ready_skus += 1
    return len(phone_ids), price_ready_skus, phone_ids


def load_phone_shop_item_ids(db_path: str | Path, shop_url: str) -> set[str]:
    connection = sqlite3.connect(db_path)
    try:
        rows = connection.execute("SELECT item_id, title FROM tmall_shop_items").fetchall()
        page_rows = connection.execute(
            "SELECT raw_json FROM tmall_shop_pages"
        ).fetchall()
    finally:
        connection.close()
    candidates = {str(item_id) for item_id, title in rows if phone_model_from_title(str(title or ""))}
    for (raw_json,) in page_rows:
        try:
            html = json.loads(raw_json)
        except (TypeError, ValueError):
            continue
        if not isinstance(html, str):
            continue
        for match in re.finditer(
            r'<dl\b[^>]*\bdata-id\s*=\s*(["\'])(?P<item_id>[^"\']+)\1[^>]*>(?P<body>.*?)</dl>',
            html,
            re.IGNORECASE | re.DOTALL,
        ):
            item_text = re.sub(r"<[^>]+>", " ", match.group("body"))
            if phone_model_from_title(item_text):
                candidates.add(match.group("item_id"))
    return candidates


def _header_columns(worksheet) -> dict[str, int]:
    return {str(cell.value): cell.column for cell in worksheet[1] if cell.value is not None}


def _exact_match(details: list[dict[str, Any]], model: str, configuration: str, color: str) -> SkuMatch | None:
    for detail in details:
        match = match_sku(detail, model, configuration, color)
        if match:
            return match
    return None


def write_backfill_workbooks(
    source_path: str | Path,
    output_path: str | Path,
    report_path: str | Path,
    details: list[dict[str, Any]],
    candidates_by_model: dict[str, list[str]] | None = None,
) -> dict[str, int]:
    """Copy the workbook, fill exact matches, and emit one report per gap."""
    book = openpyxl.load_workbook(source_path)
    sheet = book.active
    columns = _header_columns(sheet)
    required = ["机型(*)", "配置(*)", "颜色(*)", "价格", "免息分期"]
    missing = [header for header in required if header not in columns]
    if missing:
        raise ValueError(f"workbook missing columns: {', '.join(missing)}")

    report_book = openpyxl.Workbook()
    report_sheet = report_book.active
    report_sheet.title = "未匹配报告"
    report_sheet.append(REPORT_HEADERS)
    filled_prices = 0
    report_rows = 0
    candidates_by_model = candidates_by_model or {}

    for row_number in range(2, sheet.max_row + 1):
        model = str(sheet.cell(row_number, columns["机型(*)"]).value or "")
        configuration = str(sheet.cell(row_number, columns["配置(*)"]).value or "")
        color = str(sheet.cell(row_number, columns["颜色(*)"]).value or "")
        match = _exact_match(details, model, configuration, color)
        candidate_ids = candidates_by_model.get(model, [])
        if not match:
            reason = "店铺列表未找到机型" if not candidate_ids else "无精确SKU匹配"
            report_sheet.append([row_number, model, configuration, color, ",".join(candidate_ids), "未填充", "未填充", reason])
            report_rows += 1
            continue

        sheet.cell(row_number, columns["价格"]).value = match.price
        filled_prices += 1
        installment = format_installment(match.price, match.installment_periods)
        if installment:
            sheet.cell(row_number, columns["免息分期"]).value = installment
        else:
            report_sheet.append([
                row_number,
                model,
                configuration,
                color,
                match.item_id,
                "已填充",
                "未填充",
                "未提供免息分期",
            ])
            report_rows += 1

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    Path(report_path).parent.mkdir(parents=True, exist_ok=True)
    book.save(output_path)
    report_book.save(report_path)
    return {"filled_prices": filled_prices, "report_rows": report_rows}


def _successful_detail_ids(db_path: str | Path) -> set[str]:
    connection = sqlite3.connect(db_path)
    try:
        return {
            str(row[0])
            for row in connection.execute(
                "SELECT item_id FROM taobao_item_details WHERE status = 'success'"
            )
        }
    finally:
        connection.close()


def supplement_details(
    models: list[str],
    shop_db: str | Path,
    detail_db: str | Path,
    shop_url: str,
    crawler: Any,
) -> SupplementResult:
    """Run one candidate at a time and pause immediately after its first error."""
    successful_ids = _successful_detail_ids(detail_db) if Path(detail_db).exists() else set()
    attempted_ids: list[str] = []
    seen_ids: set[str] = set()
    for model in models:
        for item_id in load_shop_candidates(shop_db, shop_url, model):
            if item_id in seen_ids or item_id in successful_ids:
                continue
            seen_ids.add(item_id)
            attempted_ids.append(item_id)
            if crawler(item_id) != 0:
                return SupplementResult(attempted_ids, True)
            successful_ids.add(item_id)
    return SupplementResult(attempted_ids, False)


DEFAULT_SOURCE = Path("target/测试sku表格.xlsx")
DEFAULT_OUTPUT = Path("target/测试sku表格_已填充.xlsx")
DEFAULT_REPORT = Path("target/测试sku表格_未匹配报告.xlsx")
DEFAULT_PHONE_EXPORT = Path("target/iQOO店铺手机SKU信息.xlsx")
DEFAULT_DETAIL_DB = Path("data/taobao_items.sqlite3")
DEFAULT_SHOP_DB = Path("data/taobao_shop_items.sqlite3")
DEFAULT_SHOP_URL = "https://iqoo.tmall.com/search.htm?orderType=defaultSort&viewType=grid"


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Fill iQOO exact SKU prices into an XLSX workbook.")
    parser.add_argument("--source", default=str(DEFAULT_SOURCE))
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    parser.add_argument("--report", default=str(DEFAULT_REPORT))
    parser.add_argument("--detail-db", default=str(DEFAULT_DETAIL_DB))
    parser.add_argument("--shop-db", default=str(DEFAULT_SHOP_DB))
    parser.add_argument("--shop-url", default=DEFAULT_SHOP_URL)
    parser.add_argument("--no-crawl", action="store_true", help="Do not supplement missing local details.")
    parser.add_argument("--phone-export", help="Write all captured iQOO phone SKUs to this XLSX path.")
    return parser.parse_args(argv)


def workbook_models(source_path: str | Path) -> list[str]:
    book = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
    sheet = book.active
    columns = _header_columns(sheet)
    model_column = columns["机型(*)"]
    return list(dict.fromkeys(str(sheet.cell(row, model_column).value or "") for row in range(2, sheet.max_row + 1)))


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    if args.phone_export:
        rows = build_phone_export_rows(load_detail_records(args.detail_db), args.shop_url)
        written = write_phone_export(args.phone_export, rows)
        print(f"Phone SKU export rows={written}")
        return 0
    models = workbook_models(args.source)
    candidates_by_model = {
        model: load_shop_candidates(args.shop_db, args.shop_url, model) for model in models
    }
    if not args.no_crawl:
        try:
            from src import taobao_batch
        except ModuleNotFoundError:
            import taobao_batch

        def crawl_one(item_id: str) -> int:
            batch_args = taobao_batch.parse_args(
                [
                    "--ids", item_id,
                    "--db", args.detail_db,
                    "--output-dir", "",
                    "--delay-min", "8",
                    "--delay-max", "15",
                ]
            )
            return taobao_batch.crawl_batch(batch_args)

        supplement = supplement_details(
            models, args.shop_db, args.detail_db, args.shop_url, crawl_one
        )
        if supplement.paused:
            print(f"Supplementary crawl paused after item_id={supplement.attempted_ids[-1]}")
    result = write_backfill_workbooks(
        args.source,
        args.output,
        args.report,
        load_detail_records(args.detail_db),
        candidates_by_model,
    )
    print(f"Filled prices={result['filled_prices']} report_rows={result['report_rows']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
