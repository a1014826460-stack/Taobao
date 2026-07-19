import argparse
import json
import re
import sqlite3
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


CHINESE_COLUMNS = [
    "平台名称",
    "商品ID",
    "spu名称",
    "skuid",
    "sku描述",
    "正常价格（标价）",
    "到手价",
    "销售状态",
    "销量",
    "商品链接",
    "店铺名称",
    "店铺文本ID",
    "店铺链接",
    "收录日期",
    "核查时间",
    "发货地区",
    "优惠信息",
]

DEFAULT_DB = Path("data") / "jd_ware_business_details.sqlite3"
DEFAULT_OUTPUT = Path("data") / "jd_ware_business_details.xlsx"


def utc_now_iso():
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def parse_num_iids(values):
    seen = set()
    result = []
    for value in values or []:
        for part in re.split(r"[\s,]+", str(value).strip()):
            num_iid = part.strip().lstrip("\ufeff")
            if not num_iid or num_iid in seen:
                continue
            seen.add(num_iid)
            result.append(num_iid)
    return result


class SQLiteJDWareBusinessStore:
    def __init__(self, db_path=DEFAULT_DB):
        self.db_path = str(db_path)
        Path(self.db_path).parent.mkdir(parents=True, exist_ok=True)
        self.conn = sqlite3.connect(self.db_path)
        self.conn.row_factory = sqlite3.Row
        self._init_schema()

    def close(self):
        self.conn.close()

    def _init_schema(self):
        self.conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS jd_ware_business_details (
                num_iid TEXT PRIMARY KEY,
                title TEXT,
                price TEXT,
                original_price TEXT,
                final_price_label TEXT,
                shop_name TEXT,
                shop_id TEXT,
                vender_id TEXT,
                detail_url TEXT,
                stock_status TEXT,
                delivery_area TEXT,
                sales TEXT,
                api_url TEXT,
                http_status INTEGER,
                raw_json TEXT NOT NULL,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS jd_ware_business_state (
                num_iid TEXT PRIMARY KEY,
                status TEXT NOT NULL,
                http_status INTEGER,
                last_error TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );
            """
        )
        self.conn.commit()

    def get_state(self, num_iid):
        row = self.conn.execute(
            "SELECT * FROM jd_ware_business_state WHERE num_iid = ?",
            (str(num_iid),),
        ).fetchone()
        return dict(row) if row else None

    def mark_pending(self, num_iid):
        self._upsert_state(num_iid, "pending")

    def mark_error(self, num_iid, error, http_status=None):
        self._upsert_state(num_iid, "error", str(error), http_status)

    def save_success(self, num_iid, response, api_url="", http_status=200):
        now = utc_now_iso()
        row = build_export_row(str(num_iid), response, now)
        with self.conn:
            self.conn.execute(
                """
                INSERT INTO jd_ware_business_details (
                    num_iid, title, price, original_price, final_price_label,
                    shop_name, shop_id, vender_id, detail_url, stock_status,
                    delivery_area, sales, api_url, http_status, raw_json,
                    created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(num_iid) DO UPDATE SET
                    title = excluded.title,
                    price = excluded.price,
                    original_price = excluded.original_price,
                    final_price_label = excluded.final_price_label,
                    shop_name = excluded.shop_name,
                    shop_id = excluded.shop_id,
                    vender_id = excluded.vender_id,
                    detail_url = excluded.detail_url,
                    stock_status = excluded.stock_status,
                    delivery_area = excluded.delivery_area,
                    sales = excluded.sales,
                    api_url = excluded.api_url,
                    http_status = excluded.http_status,
                    raw_json = excluded.raw_json,
                    updated_at = excluded.updated_at
                """,
                (
                    str(num_iid),
                    row["spu名称"],
                    row["到手价"],
                    row["正常价格（标价）"],
                    response.get("price", {}).get("finalPrice", {}).get("priceContent", "") if isinstance(response.get("price"), dict) else "",
                    row["店铺名称"],
                    row["店铺文本ID"],
                    text((response.get("itemShopInfo") or {}).get("venderId")) if isinstance(response.get("itemShopInfo"), dict) else "",
                    row["商品链接"],
                    row["销售状态"],
                    row["发货地区"],
                    row["销量"],
                    api_url,
                    int(http_status or 0),
                    json.dumps(response, ensure_ascii=False),
                    now,
                    now,
                ),
            )
            self._upsert_state(num_iid, "success", None, http_status)

    def _upsert_state(self, num_iid, status, last_error=None, http_status=None):
        now = utc_now_iso()
        with self.conn:
            self.conn.execute(
                """
                INSERT INTO jd_ware_business_state (
                    num_iid, status, http_status, last_error, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?)
                ON CONFLICT(num_iid) DO UPDATE SET
                    status = excluded.status,
                    http_status = excluded.http_status,
                    last_error = excluded.last_error,
                    updated_at = excluded.updated_at
                """,
                (str(num_iid), status, http_status, last_error, now, now),
            )


def build_export_row(num_iid, response, checked_at):
    sku_id = str(num_iid)
    price = response.get("price") if isinstance(response, dict) else {}
    if not isinstance(price, dict):
        price = {}
    final_price = price.get("finalPrice") if isinstance(price.get("finalPrice"), dict) else {}
    shop = response.get("itemShopInfo") if isinstance(response, dict) else {}
    if not isinstance(shop, dict):
        shop = {}
    stock = response.get("stockVO") if isinstance(response, dict) else {}
    if not isinstance(stock, dict):
        stock = {}
    sku_head = response.get("skuHeadVO") if isinstance(response, dict) else {}
    if not isinstance(sku_head, dict):
        sku_head = {}
    ware_info = response.get("wareInfo") if isinstance(response, dict) else {}
    if not isinstance(ware_info, dict):
        ware_info = {}
    return {
        "平台名称": "京东",
        "商品ID": sku_id,
        "spu名称": text(first_present(sku_head.get("skuName"), sku_head.get("name"), response.get("title"))),
        "skuid": sku_id,
        "sku描述": text(first_present(sku_head.get("skuName"), sku_head.get("skuTitle"))),
        "正常价格（标价）": text(first_present(price.get("p"), price.get("op"), price.get("m"))),
        "到手价": text(first_present(final_price.get("price"), price.get("p"))),
        "销售状态": sales_status(response),
        "销量": text(first_present(response.get("commentNoticeVO", {}).get("commentCount") if isinstance(response.get("commentNoticeVO"), dict) else "", response.get("sales"))),
        "商品链接": f"https://item.jd.com/{sku_id}.html",
        "店铺名称": text(first_present(shop.get("shopName"), shop.get("name"))),
        "店铺文本ID": text(first_present(shop.get("shopId"), shop.get("venderId"))),
        "店铺链接": shop_url(first_present(shop.get("shopId"), shop.get("venderId"))),
        "收录日期": text(checked_at),
        "核查时间": text(checked_at),
        "发货地区": text(first_present(stock.get("areaName"), stock.get("sendAddr"), response.get("ipCityCode"))),
        "优惠信息": discount_info(response),
    }


def sales_status(response):
    stock = response.get("stockVO") if isinstance(response, dict) else {}
    if not isinstance(stock, dict):
        stock = {}
    desc = first_present(stock.get("stockStateDesc"), stock.get("stockDesc"))
    if desc:
        return text(desc)
    ware_map = ((response.get("wareInfo") or {}).get("wareInfoMap") or {}) if isinstance(response.get("wareInfo"), dict) else {}
    if str(ware_map.get("sku_status", "")) == "1":
        return "在售"
    return ""


def discount_info(response):
    values = []
    promotion = response.get("promotion") if isinstance(response, dict) else None
    best = response.get("bestPromotion") if isinstance(response, dict) else None
    for value in (best, promotion):
        if value in (None, "", [], {}):
            continue
        if isinstance(value, (dict, list)):
            values.append(json.dumps(value, ensure_ascii=False))
        else:
            values.append(str(value))
    return "; ".join(values)


def shop_url(shop_id):
    if not shop_id:
        return ""
    return f"https://mall.jd.com/index-{shop_id}.html"


def first_present(*values):
    for value in values:
        if value is not None and value != "":
            return value
    return ""


def text(value):
    if value is None:
        return ""
    return str(value)


def flatten_json(value, prefix):
    flattened = {}
    if isinstance(value, dict):
        if not value:
            flattened[prefix] = ""
        for key, child in value.items():
            flattened.update(flatten_json(child, f"{prefix}.{key}" if prefix else str(key)))
    elif isinstance(value, list):
        if not value:
            flattened[prefix] = ""
        for index, child in enumerate(value):
            flattened.update(flatten_json(child, f"{prefix}[{index}]"))
    else:
        flattened[prefix] = scalar_to_cell(value)
    return flattened


def scalar_to_cell(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value)


def load_rows(db_path, num_iids=None):
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    try:
        if num_iids:
            placeholders = ",".join("?" for _ in num_iids)
            sql = f"SELECT * FROM jd_ware_business_details WHERE num_iid IN ({placeholders}) ORDER BY CASE num_iid"
            order_params = []
            for index, num_iid in enumerate(num_iids):
                sql += " WHEN ? THEN ?"
                order_params.extend([str(num_iid), index])
            sql += " END"
            params = [str(num_iid) for num_iid in num_iids] + order_params
            return [dict(row) for row in conn.execute(sql, params).fetchall()]
        return [dict(row) for row in conn.execute("SELECT * FROM jd_ware_business_details ORDER BY updated_at, num_iid").fetchall()]
    finally:
        conn.close()


def export_to_xlsx(db_path=DEFAULT_DB, output_path=DEFAULT_OUTPUT, num_iids=None):
    db_rows = load_rows(db_path, num_iids)
    export_rows = []
    raw_headers = []
    seen = set()
    for db_row in db_rows:
        response = json.loads(db_row["raw_json"])
        row = build_export_row(db_row["num_iid"], response, db_row.get("updated_at") or "")
        flat = flatten_json(response, "raw_json")
        for key in flat:
            if key not in seen:
                seen.add(key)
                raw_headers.append(key)
        export_rows.append({**row, **flat})

    headers = CHINESE_COLUMNS + raw_headers
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "京东商品信息"
    worksheet.append(headers)
    for row in export_rows:
        worksheet.append([row.get(column, "") for column in headers])
    style_sheet(worksheet)
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return len(export_rows)


def style_sheet(worksheet):
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.freeze_panes = "A2"
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells[:200]:
            if cell.value is None:
                continue
            length = sum(2 if ord(char) > 127 else 1 for char in str(cell.value))
            max_length = max(max_length, length)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 60)


def build_arg_parser():
    parser = argparse.ArgumentParser(description="Export JD wareBusiness SQLite rows to XLSX.")
    parser.add_argument("--db", default=str(DEFAULT_DB))
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    parser.add_argument("--num-iids", action="append", default=[])
    parser.add_argument("--num-iids-file", action="append", default=[])
    return parser


def main(argv=None):
    parser = build_arg_parser()
    args = parser.parse_args(argv)
    raw_values = list(args.num_iids or [])
    for file_path in args.num_iids_file or []:
        raw_values.append(Path(file_path).read_text(encoding="utf-8"))
    num_iids = parse_num_iids(raw_values)
    written = export_to_xlsx(args.db, args.output, num_iids or None)
    print(f"Done: {args.output} rows={written}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
