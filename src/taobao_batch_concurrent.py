"""
Concurrent batch crawler for Taobao items.

Reads product links from an xlsx file (link column), extracts the num_iid,
fetches item details via the fan-b.com gateway concurrently, and persists
results to SQLite.

Usage:
    python -m src.taobao_batch_concurrent                          # default xlsx & db
    python -m src.taobao_batch_concurrent --xlsx path.xlsx         # custom xlsx
    python -m src.taobao_batch_concurrent --workers 10             # concurrent workers
    python -m src.taobao_batch_concurrent --reset                  # re-crawl already-successful items
"""

from __future__ import annotations

import argparse
import os
import re
import sys
import time
from concurrent.futures import Future, ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterator

# --- Project root -----------------------------------------------------------
PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


# ---------------------------------------------------------------------------
# xlsx helpers
# ---------------------------------------------------------------------------

def _load_workbook(path: str):
    """Lazy-import openpyxl so it is only required when reading xlsx files."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("openpyxl is required. Install with: pip install openpyxl")
    return load_workbook(path, read_only=True, data_only=True)


def extract_num_iids_from_xlsx(xlsx_path: str, link_column: str = "link") -> list[str]:
    """Return deduplicated num_iid strings extracted from a link column.

    Links are expected to contain ``id=<num_iid>``, e.g.
        https://item.taobao.com/item.htm?id=1038508552841
    """
    wb = _load_workbook(xlsx_path)
    ws = wb.active

    # Locate header row
    headers: list[str] = []
    col_idx: int | None = None
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = [str(c).strip().lower() if c is not None else "" for c in row]
        try:
            col_idx = headers.index(link_column.lower())
        except ValueError:
            pass
        break

    if col_idx is None:
        raise ValueError(
            f"Column '{link_column}' not found in sheet '{ws.title}'. "
            f"Available columns: {headers}"
        )

    seen: set[str] = set()
    result: list[str] = []
    id_pattern = re.compile(r"id=(\d+)")

    for row in ws.iter_rows(min_row=2, values_only=True):
        link = row[col_idx] if col_idx < len(row) else None  # type: ignore[assignment]
        if link is None:
            continue
        match = id_pattern.search(str(link))
        if match:
            num_iid = match.group(1)
            if num_iid not in seen:
                seen.add(num_iid)
                result.append(num_iid)

    wb.close()
    return result


# ---------------------------------------------------------------------------
# Load credentials from environment / .env
# ---------------------------------------------------------------------------

def _load_dotenv() -> None:
    """Load .env from project root if it exists (simple parser, no dependency)."""
    env_path = PROJECT_ROOT / ".env"
    if not env_path.exists():
        return
    for line in env_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, _, value = line.partition("=")
        key, value = key.strip(), value.strip().strip("\"'")
        if key and key not in os.environ:
            os.environ[key] = value


# ---------------------------------------------------------------------------
# Concurrent crawler (reuses the proven fetch_item_detail & SQLiteItemStore)
# ---------------------------------------------------------------------------

@dataclass
class ConcurrentCrawlerConfig:
    key: str
    secret: str
    num_iids: list[str]
    db_path: str = "data/taobao_batch.sqlite3"
    reset_items: bool = False
    lang: str = "zh-CN"
    delay: float = 1.0           # per-request delay between API calls (seconds)
    timeout: float = 30.0
    retries: int = 5             # more retries with backoff
    item_api: str = "item_get"
    is_promotion: str | None = None
    max_workers: int = 4         # fewer workers to avoid rate-limiting
    progress_interval: int = 50


@dataclass
class CrawlProgress:
    total: int
    fetched: int = 0
    skipped: int = 0
    failed: int = 0
    _lock: Any = field(default_factory=lambda: __import__("threading").Lock())

    def snapshot(self) -> dict[str, int]:
        with self._lock:
            return {"total": self.total, "fetched": self.fetched,
                    "skipped": self.skipped, "failed": self.failed}


def crawl_items_concurrent(
    config: ConcurrentCrawlerConfig,
) -> CrawlProgress:
    """Crawl *num_iids* concurrently using a thread pool with rate limiting."""
    import threading
    from src.taobao.direct.item import ItemCrawlerConfig, SQLiteItemStore, fetch_item_detail  # noqa: E402

    # --------------- pre-flight: skip already-successful items ---------------
    preflight_store = SQLiteItemStore(config.db_path)
    try:
        pending_ids: list[str] = []
        skipped_count = 0
        for nid in config.num_iids:
            state = preflight_store.get_item_state(nid)
            if state and state.get("status") == "success" and not config.reset_items:
                skipped_count += 1
            else:
                pending_ids.append(nid)
    finally:
        preflight_store.close()

    progress = CrawlProgress(total=len(config.num_iids), skipped=skipped_count)

    if not pending_ids:
        print(f"All {progress.total} items already fetched - nothing to do.")
        return progress

    print(f"Total: {progress.total}  Pending (incl. errors): {len(pending_ids)}  "
          f"Skipped (already OK): {skipped_count}")
    print(f"Workers: {config.max_workers}  Delay: {config.delay}s  "
          f"Retries: {config.retries}  API: {config.item_api}")
    print("-" * 60)

    # --------------- rate limiter: token bucket via semaphore + delay ---------------
    rate_sem = threading.Semaphore(config.max_workers)
    last_request_lock = threading.Lock()
    last_request_time: float = 0.0

    def _throttle() -> None:
        """Ensure minimum interval between API requests across all workers."""
        nonlocal last_request_time
        with last_request_lock:
            now = time.monotonic()
            wait = last_request_time + config.delay - now
            if wait > 0:
                time.sleep(wait)
            last_request_time = time.monotonic()

    # --------------- per-worker config ---------------
    worker_config = ItemCrawlerConfig(
        key=config.key,
        secret=config.secret,
        num_iids=[],
        db_path=config.db_path,
        reset_items=config.reset_items,
        lang=config.lang,
        delay=0,                  # we handle throttling ourselves
        timeout=config.timeout,
        retries=config.retries,
        item_api=config.item_api,
        is_promotion=config.is_promotion,
    )

    # --------------- progress reporting ---------------
    completed_count = 0
    start_time = time.monotonic()

    def _report():
        nonlocal completed_count
        completed_count += 1
        if completed_count % config.progress_interval == 0:
            elapsed = time.monotonic() - start_time
            rate = completed_count / elapsed if elapsed > 0 else 0
            print(f"\n  [{completed_count}/{len(pending_ids)}] "
                  f"fetched={progress.fetched} failed={progress.failed} "
                  f"| {rate:.1f} items/s")

    # --------------- worker ---------------
    def _crawl_one(num_iid: str) -> None:
        """Fetch a single item with retries and persist it."""
        from src.taobao.direct.item import SQLiteItemStore, parse_item_response  # noqa: E402

        store = SQLiteItemStore(config.db_path)
        try:
            rate_sem.acquire()
            try:
                _throttle()
                store.mark_pending(num_iid)
                response = fetch_item_detail(worker_config, num_iid)
                parse_item_response(response)
                store.save_item_detail(num_iid, response)
                with progress._lock:
                    progress.fetched += 1
            finally:
                rate_sem.release()
        except Exception as exc:
            store.mark_error(num_iid, exc)
            with progress._lock:
                progress.failed += 1
        finally:
            store.close()
            _report()

    # --------------- concurrent execution ---------------
    print(f"Starting crawl at {utc_now_iso()}\n")
    with ThreadPoolExecutor(max_workers=config.max_workers) as executor:
        future_map: dict[Future[None], str] = {}
        for num_iid in pending_ids:
            future = executor.submit(_crawl_one, num_iid)
            future_map[future] = num_iid

        for future in as_completed(future_map):
            num_iid = future_map[future]
            try:
                future.result()
            except Exception as exc:
                with progress._lock:
                    progress.failed += 1
                print(f"\n  !! Unhandled error for {num_iid}: {exc}")

    elapsed = time.monotonic() - start_time
    print(f"\n{'=' * 60}")
    print(f"Finished at {utc_now_iso()}")
    print(f"Total: {progress.total}  Fetched: {progress.fetched}  "
          f"Skipped: {progress.skipped}  Failed: {progress.failed}")
    if elapsed > 0:
        print(f"Duration: {elapsed:.1f}s  Rate: {progress.fetched / elapsed:.2f} items/s")
    return progress


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Concurrent batch Taobao item crawler",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--xlsx",
        default="target/products_20260726.xlsx",
        help="Path to xlsx with a 'link' column (default: target/products_20260726.xlsx)",
    )
    parser.add_argument(
        "--link-column",
        default="link",
        help="Column name containing the product links (default: link)",
    )
    parser.add_argument(
        "--db",
        default="data/taobao_batch.sqlite3",
        help="SQLite database path (default: data/taobao_batch.sqlite3)",
    )
    parser.add_argument(
        "--workers",
        type=int,
        default=4,
        help="Number of concurrent workers (default: 4)",
    )
    parser.add_argument(
        "--api",
        default="item_get",
        choices=["item_get", "item_get_pro"],
        help="Gateway API endpoint: item_get or item_get_pro (default: item_get)",
    )
    parser.add_argument(
        "--delay",
        type=float,
        default=1.0,
        help="Per-request minimum interval in seconds (default: 1.0)",
    )
    parser.add_argument(
        "--timeout",
        type=float,
        default=30.0,
        help="HTTP request timeout in seconds (default: 30.0)",
    )
    parser.add_argument(
        "--retries",
        type=int,
        default=5,
        help="Max retries per item (default: 5)",
    )
    parser.add_argument(
        "--reset",
        action="store_true",
        help="Re-crawl items that were already fetched successfully",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Only extract IDs and print summary, do not crawl",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = _parse_args(argv)

    # Resolve paths relative to project root
    xlsx_path: str = str(PROJECT_ROOT / args.xlsx) if not os.path.isabs(args.xlsx) else args.xlsx
    db_path: str = str(PROJECT_ROOT / args.db) if not os.path.isabs(args.db) else args.db

    # 1. Extract IDs
    print(f"Reading links from: {xlsx_path}")
    num_iids = extract_num_iids_from_xlsx(xlsx_path, args.link_column)
    print(f"Extracted {len(num_iids)} unique product IDs")

    if not num_iids:
        print("No product IDs found — exiting.")
        return 1

    if args.dry_run:
        print("Dry-run mode — skipping crawl.")
        print(f"First 5 IDs: {num_iids[:5]}")
        print(f"Last 5 IDs:  {num_iids[-5:]}")
        return 0

    # 2. Load credentials
    _load_dotenv()
    # Accept multiple naming conventions: FANB_API_KEY / KEY, FANB_API_SECRET / SECRET
    key = os.environ.get("FANB_API_KEY") or os.environ.get("KEY") or ""
    secret = os.environ.get("FANB_API_SECRET") or os.environ.get("SECRET") or ""

    if not key or not secret:
        print("ERROR: API credentials must be set in .env file.")
        print("  Use either of these naming conventions:")
        print("    FANB_API_KEY=...    /  KEY=...")
        print("    FANB_API_SECRET=... /  SECRET=...")
        return 2

    # 3. Crawl
    config = ConcurrentCrawlerConfig(
        key=key,
        secret=secret,
        num_iids=num_iids,
        db_path=db_path,
        reset_items=args.reset,
        delay=args.delay,
        timeout=args.timeout,
        retries=args.retries,
        item_api=args.api,
        max_workers=args.workers,
    )

    progress = crawl_items_concurrent(config)

    if progress.failed > 0:
        print(f"\nWARNING: {progress.failed} items failed. "
              f"Re-run with the same --db to retry only failed items.")
    return 0 if progress.failed == 0 else 3


if __name__ == "__main__":
    raise SystemExit(main())
