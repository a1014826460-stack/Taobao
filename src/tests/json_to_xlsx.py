"""
将 data/sample.json 转换为 xlsx 文件。

JSON 结构为嵌套的淘宝商品数据，本脚本将其拆分为多个 Sheet：
  - Sheet "商品概要": item 的平铺字段（不含嵌套列表）
  - Sheet "SKU列表": skus.sku 数组
  - Sheet "商品属性": props 数组
  - Sheet "商品图片": item_imgs 数组
  - Sheet "详情图片": desc_img 数组
"""

import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def flatten_item(item: dict) -> dict:
    """提取 item 中非嵌套的基础字段，列表/字典字段统一转为字符串或省略。"""
    skip_keys = {
        "desc_img", "item_imgs", "skus", "props", "props_list",
        "props_imgs", "props_img", "seller_info", "video",
        "crumbs", "url_log", "language", "call_args",
    }
    row = {}
    for k, v in item.items():
        if k in skip_keys:
            continue
        if isinstance(v, (list, dict)):
            row[k] = json.dumps(v, ensure_ascii=False)
        elif v is None:
            row[k] = ""
        else:
            row[k] = v
    # 把 seller_info 的子字段平铺进来
    seller = item.get("seller_info", {})
    if isinstance(seller, dict):
        for sk, sv in seller.items():
            row[f"seller_{sk}"] = sv if sv is not None else ""
    return row


def style_header(ws, headers: list):
    """给表头加样式。"""
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")


def auto_width(ws):
    """根据内容自动调整列宽（取前 100 行做参考）。"""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells[:100]:
            if cell.value:
                # 中文字符按 2 个字符宽度计算
                val = str(cell.value)
                length = sum(2 if ord(c) > 127 else 1 for c in val)
                max_len = max(max_len, length)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)


def main():
    base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    json_path = os.path.join(base_dir, "data", "sample.json")
    xlsx_path = os.path.join(base_dir, "data", "sample.xlsx")

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    item = data.get("item", {})

    wb = Workbook()

    # ==================== Sheet 1: 商品概要 ====================
    ws1 = wb.active
    ws1.title = "商品概要"
    flat = flatten_item(item)
    # 转置：字段名在 A 列，值在 B 列
    ws1.append(["字段", "值"])
    for k, v in flat.items():
        ws1.append([k, v])
    style_header(ws1, ["字段", "值"])
    auto_width(ws1)

    # ==================== Sheet 2: SKU 列表 ====================
    ws2 = wb.create_sheet("SKU列表")
    skus = item.get("skus", {}).get("sku", [])
    if skus:
        headers = list(skus[0].keys())
        ws2.append(headers)
        for sku in skus:
            ws2.append([sku.get(h, "") for h in headers])
        style_header(ws2, headers)
        auto_width(ws2)
    else:
        ws2.append(["无 SKU 数据"])

    # ==================== Sheet 3: 商品属性 ====================
    ws3 = wb.create_sheet("商品属性")
    props = item.get("props", [])
    if props:
        ws3.append(["属性名", "属性值"])
        for p in props:
            ws3.append([p.get("name", ""), p.get("value", "")])
        style_header(ws3, ["属性名", "属性值"])
        auto_width(ws3)
    else:
        ws3.append(["无属性数据"])

    # ==================== Sheet 4: 商品图片 ====================
    ws4 = wb.create_sheet("商品图片")
    imgs = item.get("item_imgs", [])
    if imgs:
        ws4.append(["序号", "图片URL"])
        for i, img in enumerate(imgs, 1):
            ws4.append([i, img.get("url", "")])
        style_header(ws4, ["序号", "图片URL"])
        auto_width(ws4)
    else:
        ws4.append(["无图片数据"])

    # ==================== Sheet 5: 详情图片 ====================
    ws5 = wb.create_sheet("详情图片")
    desc_imgs = item.get("desc_img", [])
    if desc_imgs:
        ws5.append(["序号", "图片URL"])
        for i, url in enumerate(desc_imgs, 1):
            ws5.append([i, url])
        style_header(ws5, ["序号", "图片URL"])
        auto_width(ws5)
    else:
        ws5.append(["无详情图片数据"])

    wb.save(xlsx_path)
    print(f"Done: {xlsx_path}")


if __name__ == "__main__":
    main()
