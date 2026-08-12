"""Export 黄小米 JD details/comments with the exact data_sample.xlsx headers."""

import argparse
import json
import sqlite3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


HEADERS = [
    "搜索关键词", "商品SKU", "商品链接", "商品名称", "到手价", "原价", "已售数", "评论数",
    "好评数", "中评数", "差评数", "追评数", "店铺ID", "店铺链接", "店铺名称", "店铺星级",
    "店铺关注/粉丝", "商品描述信息", "商品服务信息", "商品产地", "是否为京东物流", "是否包邮",
    "是否有广告标识", "是否为自营店", "小米等级", "保质期",
]


def text(value):
    return "" if value is None else str(value)


def value_at(data, *paths):
    for path in paths:
        current = data
        for key in path.split("."):
            current = current.get(key) if isinstance(current, dict) else None
        if current not in (None, "", [], {}):
            return current
    return ""


def join_values(values):
    return "|".join(dict.fromkeys(text(value).strip() for value in values if text(value).strip()))


def comment_summary(comment_raw):
    data = comment_raw.get("data") if isinstance(comment_raw, dict) else {}
    return data.get("Summary") if isinstance(data, dict) and isinstance(data.get("Summary"), dict) else {}


def comment_texts(comment_raw):
    data = comment_raw.get("data") if isinstance(comment_raw, dict) else {}
    items = data.get("items") if isinstance(data, dict) else []
    if not isinstance(items, list):
        return ""
    return join_values(value_at(item, "displayText", "text") for item in items if isinstance(item, dict))


def property_values(item):
    props = item.get("props") if isinstance(item, dict) else []
    if not isinstance(props, list):
        return []
    result = []
    for prop in props:
        if isinstance(prop, dict):
            result.append(f"{text(prop.get('name')).strip()}:{text(prop.get('value')).strip()}".strip(":"))
    return result


def detect_yes_no(text_blob, yes_words, no_words=()):
    if any(word in text_blob for word in yes_words):
        return "是"
    if any(word in text_blob for word in no_words):
        return "否"
    return ""


def build_row(detail, comment):
    raw = json.loads(detail["raw_json"])
    item = raw.get("item") if isinstance(raw.get("item"), dict) else {}
    summary = comment_summary(json.loads(comment["raw_json"])) if comment else {}
    title = text(item.get("title") or detail["title"])
    props = property_values(item)
    props_text = join_values(props)
    desc = join_values([item.get("desc_short"), item.get("desc"), comment_texts(json.loads(comment["raw_json"])) if comment else ""])
    service = join_values([item.get("post_fee"), item.get("express_fee"), item.get("ems_fee")])
    shop_id = text(item.get("shop_id") or detail["shop_id"])
    item_url = text(item.get("detail_url") or detail["detail_url"] or f"https://item.jd.com/{detail['num_iid']}.html")
    blob = "|".join([title, props_text, desc, service])
    return [
        "黄小米相关词", detail["num_iid"], item_url, title,
        text(item.get("promotion_price") or item.get("price") or detail["price"]),
        text(item.get("orginal_price") or detail["orginal_price"]),
        text(item.get("total_sold") or item.get("sales") or detail["sales"]),
        text(summary.get("CommentCountStr") or summary.get("CommentCount")),
        text(summary.get("GoodCountStr") or summary.get("GoodCount")),
        text(summary.get("GeneralCountStr") or summary.get("GeneralCount")),
        text(summary.get("PoorCountStr") or summary.get("PoorCount")),
        text(summary.get("AfterCountStr") or summary.get("AfterCount")),
        shop_id, f"https://mall.jd.com/index-{shop_id}.html" if shop_id else "", text(item.get("nick") or detail["nick"]),
        "", "", desc, service, "",
        detect_yes_no(blob, ("京东物流", "京东配送", "京东发货")),
        detect_yes_no(blob, ("包邮", "免邮")), detect_yes_no(blob, ("广告",)),
        detect_yes_no(blob, ("京东自营", "自营")),
        "", "",
    ]


def export_xlsx(detail_db, comment_db, output):
    details = sqlite3.connect(detail_db)
    details.row_factory = sqlite3.Row
    comments = sqlite3.connect(comment_db)
    comments.row_factory = sqlite3.Row
    try:
        comment_rows = {row["itemid"]: row for row in comments.execute("SELECT * FROM jd_item_comments")}
        rows = [build_row(row, comment_rows.get(row["num_iid"])) for row in details.execute("SELECT * FROM jd_item_details ORDER BY num_iid")]
    finally:
        details.close()
        comments.close()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "工作表1"
    sheet.append(HEADERS)
    for row in rows:
        sheet.append(row)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center")
    sheet.freeze_panes = "A2"
    for column in sheet.columns:
        sheet.column_dimensions[column[0].column_letter].width = min(max(12, max(len(text(cell.value)) for cell in column[:100]) + 2), 50)
    Path(output).parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return len(rows)


def main(argv=None):
    parser = argparse.ArgumentParser(description="Export 黄小米 JD data using data_sample.xlsx headers.")
    parser.add_argument("--detail-db", default="data/jd_huangxiaomi_item_details.sqlite3")
    parser.add_argument("--comment-db", default="data/jd_huangxiaomi_comments.sqlite3")
    parser.add_argument("--output", default="target/jd_huangxiaomi_20260809/黄小米商品数据.xlsx")
    args = parser.parse_args(argv)
    print(f"Done: {args.output} rows={export_xlsx(args.detail_db, args.comment_db, args.output)}")


if __name__ == "__main__":
    main()
