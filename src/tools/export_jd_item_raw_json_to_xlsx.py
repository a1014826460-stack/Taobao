import argparse
import json
import re
import sqlite3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PRIORITY_COLUMNS = [
    "平台名称",
    "商品ID",
    "spu名称",
    "skuid",
    "sku描述",
    "正常价格（标价）",
    "到手价",
    "销售状态",
    "销量",
    "商品链接",
    "店铺名称",
    "店铺文本ID",
    "店铺链接",
    "收录日期",
    "核查时间",
    "发货地区",
    "优惠信息",
]

DEFAULT_DB = Path("data") / "jd_item_details.sqlite3"
DEFAULT_OUTPUT = Path("data") / "jd_item_details.xlsx"


def parse_num_iids(values):
    seen = set()
    result = []
    for value in values or []:
        for part in re.split(r"[\s,]+", str(value).strip()):
            num_iid = part.strip().lstrip("\ufeff")
            if not num_iid or num_iid in seen:
                continue
            seen.add(num_iid)
            result.append(num_iid)
    return result


def export_jd_item_details_to_xlsx(db_path, output_path, num_iids=None):
    db_rows = load_jd_item_detail_rows(db_path, num_iids)
    rows = []
    raw_headers = []
    seen_raw_headers = set()
    for db_row in db_rows:
        response = json.loads(db_row["raw_json"])
        flat_raw = flatten_json(response, prefix="raw_json")
        for key in flat_raw:
            if key not in seen_raw_headers:
                seen_raw_headers.add(key)
                raw_headers.append(key)
        rows.append((db_row, response, flat_raw))

    headers = PRIORITY_COLUMNS + raw_headers
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "京东商品信息"
    worksheet.append(headers)

    written = 0
    for db_row, response, flat_raw in rows:
        for export_row in build_export_rows(db_row, response):
            merged_row = {**export_row, **flat_raw}
            worksheet.append([merged_row.get(column, "") for column in headers])
            written += 1

    style_sheet(worksheet)
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return written


def load_jd_item_detail_rows(db_path, num_iids=None):
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    try:
        if num_iids:
            placeholders = ",".join("?" for _ in num_iids)
            sql = f"SELECT * FROM jd_item_details WHERE num_iid IN ({placeholders}) ORDER BY CASE num_iid"
            order_params = []
            for index, num_iid in enumerate(num_iids):
                sql += " WHEN ? THEN ?"
                order_params.extend([str(num_iid), index])
            sql += " END"
            params = [str(num_iid) for num_iid in num_iids] + order_params
            return [dict(row) for row in conn.execute(sql, params).fetchall()]
        return [
            dict(row)
            for row in conn.execute(
                "SELECT * FROM jd_item_details ORDER BY updated_at, num_iid"
            ).fetchall()
        ]
    finally:
        conn.close()


def build_export_rows(db_row, response):
    item = response.get("item") or {}
    skus = ((item.get("skus") or {}).get("sku") or []) if isinstance(item.get("skus"), dict) else []
    if not skus:
        skus = [{}]
    return [build_export_row(db_row, response, item, sku) for sku in skus]


def build_export_row(db_row, response, item, sku):
    seller_info = item.get("seller_info") if isinstance(item.get("seller_info"), dict) else {}
    return {
        "平台名称": "京东",
        "商品ID": text(item.get("num_iid") or db_row.get("num_iid")),
        "spu名称": text(item.get("title") or db_row.get("title")),
        "skuid": text(sku.get("sku_id") or sku.get("skuId")),
        "sku描述": text(sku.get("properties_name") or sku.get("properties") or sku.get("name")),
        "正常价格（标价）": text(
            first_present(
                sku.get("orginal_price"),
                sku.get("original_price"),
                item.get("orginal_price"),
                item.get("original_price"),
                db_row.get("orginal_price"),
            )
        ),
        "到手价": text(first_present(sku.get("price"), item.get("price"), db_row.get("price"))),
        "销售状态": sales_status(item, sku),
        "销量": text(first_present(item.get("sales"), item.get("total_sold"), db_row.get("sales"))),
        "商品链接": text(item.get("detail_url") or db_row.get("detail_url")),
        "店铺名称": text(seller_info.get("shop_name") or item.get("nick") or db_row.get("nick")),
        "店铺文本ID": text(item.get("shop_id") or db_row.get("shop_id")),
        "店铺链接": text(seller_info.get("zhuy") or item.get("shop_url") or shop_url(item.get("shop_id") or db_row.get("shop_id"))),
        "收录日期": text(db_row.get("created_at")),
        "核查时间": text(db_row.get("updated_at")),
        "发货地区": text(item.get("location") or item.get("delivery_from")),
        "优惠信息": discount_info(item, sku),
    }


def flatten_json(value, prefix):
    flattened = {}
    if isinstance(value, dict):
        if not value:
            flattened[prefix] = ""
        for key, child in value.items():
            flattened.update(flatten_json(child, f"{prefix}.{key}" if prefix else str(key)))
    elif isinstance(value, list):
        if not value:
            flattened[prefix] = ""
        for index, child in enumerate(value):
            flattened.update(flatten_json(child, f"{prefix}[{index}]"))
    else:
        flattened[prefix] = scalar_to_cell(value)
    return flattened


def scalar_to_cell(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value)


def first_present(*values):
    for value in values:
        if value is not None and value != "":
            return value
    return ""


def text(value):
    if value is None:
        return ""
    return str(value)


def sales_status(item, sku):
    quantity = first_present(sku.get("quantity"), item.get("num"), item.get("stock"))
    if quantity == "":
        return ""
    try:
        return "在售" if int(float(quantity)) > 0 else "无库存"
    except (TypeError, ValueError):
        return ""


def shop_url(shop_id):
    if not shop_id:
        return ""
    return f"https://mall.jd.com/index-{shop_id}.html"


def discount_info(item, sku):
    values = []
    for key in ("promotion_price", "coupon_info", "discount", "has_discount"):
        value = first_present(sku.get(key), item.get(key))
        if value not in {"", None, False}:
            values.append(f"{key}:{value}")
    return "; ".join(values)


def style_sheet(worksheet):
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.freeze_panes = "A2"
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells[:200]:
            if cell.value is None:
                continue
            length = sum(2 if ord(char) > 127 else 1 for char in str(cell.value))
            max_length = max(max_length, length)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 60)


def build_arg_parser():
    parser = argparse.ArgumentParser(
        description="Export JD item raw_json rows from SQLite to an XLSX file."
    )
    parser.add_argument("--db", default=str(DEFAULT_DB), help="SQLite database path.")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT), help="Output XLSX path.")
    parser.add_argument(
        "--num-iids",
        action="append",
        default=[],
        help="Specific JD item IDs. Supports comma, whitespace, and repeated arguments.",
    )
    parser.add_argument(
        "--num-iids-file",
        action="append",
        default=[],
        help="Text file containing JD item IDs separated by comma, whitespace, or newline.",
    )
    return parser


def main(argv=None):
    parser = build_arg_parser()
    args = parser.parse_args(argv)
    raw_values = list(args.num_iids or [])
    for file_path in args.num_iids_file or []:
        raw_values.append(Path(file_path).read_text(encoding="utf-8"))
    num_iids = parse_num_iids(raw_values)
    written = export_jd_item_details_to_xlsx(args.db, args.output, num_iids or None)
    print(f"Done: {args.output} rows={written}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
