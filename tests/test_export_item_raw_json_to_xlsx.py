import json
import os
import sqlite3
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from src.tools.export_item_raw_json_to_xlsx import (
    PRIORITY_COLUMNS,
    export_item_details_to_xlsx,
    parse_num_iids,
)


class ExportItemRawJsonToXlsxTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.db_path = os.path.join(self.tmp.name, "items.sqlite3")
        self.xlsx_path = os.path.join(self.tmp.name, "items.xlsx")
        conn = sqlite3.connect(self.db_path)
        conn.execute(
            """
            CREATE TABLE item_details (
                num_iid TEXT PRIMARY KEY,
                title TEXT,
                price TEXT,
                orginal_price TEXT,
                nick TEXT,
                detail_url TEXT,
                pic_url TEXT,
                brand TEXT,
                cid TEXT,
                seller_id TEXT,
                shop_id TEXT,
                sales TEXT,
                raw_json TEXT NOT NULL,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        raw_json = {
            "api_type": "taobao",
            "server_time": "Beijing/2026-06-26 09:34:59",
            "item": {
                "num_iid": "599262347474",
                "title": "sample title",
                "price": "67.00",
                "orginal_price": "88.00",
                "sales": 2,
                "detail_url": "https://item.taobao.com/item.htm?id=599262347474",
                "nick": "demo shop",
                "shop_id": "486527013",
                "seller_id": "3013891076",
                "location": "广东江门",
                "seller_info": {
                    "shop_name": "demo shop name",
                    "zhuy": "https://shop486527013.taobao.com/",
                },
                "props": [
                    {
                        "name": "品牌",
                        "value": "demo brand",
                    }
                ],
                "skus": {
                    "sku": [
                        {
                            "sku_id": 4178686389305,
                            "properties_name": "口味:冬阴功",
                            "orginal_price": 88,
                            "price": 67,
                            "quantity": "200",
                        }
                    ]
                },
            },
        }
        conn.execute(
            """
            INSERT INTO item_details (
                num_iid, title, price, orginal_price, nick, detail_url,
                pic_url, brand, cid, seller_id, shop_id, sales, raw_json,
                created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "599262347474",
                "sample title",
                "67.00",
                "88.00",
                "demo shop",
                "https://item.taobao.com/item.htm?id=599262347474",
                "",
                "",
                "",
                "3013891076",
                "486527013",
                "2",
                json.dumps(raw_json, ensure_ascii=False),
                "2026-06-26T01:00:00+00:00",
                "2026-06-26T02:00:00+00:00",
            ),
        )
        conn.commit()
        conn.close()

    def tearDown(self):
        self.tmp.cleanup()

    def test_parse_num_iids_deduplicates_multiple_input_formats(self):
        values = parse_num_iids(["1, 2", "2\n3"])

        self.assertEqual(values, ["1", "2", "3"])

    def test_export_writes_priority_columns_first_and_expands_skus(self):
        written = export_item_details_to_xlsx(
            db_path=self.db_path,
            output_path=self.xlsx_path,
            num_iids=["599262347474"],
        )

        self.assertEqual(written, 1)
        wb = load_workbook(self.xlsx_path)
        ws = wb["商品信息"]
        headers = [cell.value for cell in ws[1]][: len(PRIORITY_COLUMNS)]
        row = [cell.value for cell in ws[2]][: len(PRIORITY_COLUMNS)]

        self.assertEqual(headers, PRIORITY_COLUMNS)
        self.assertEqual(row[0], "淘宝")
        self.assertEqual(row[1], "599262347474")
        self.assertEqual(row[2], "sample title")
        self.assertEqual(row[3], "4178686389305")
        self.assertEqual(row[4], "口味:冬阴功")
        self.assertEqual(row[5], "88")
        self.assertEqual(row[6], "67")
        self.assertEqual(row[8], "2")
        self.assertEqual(row[10], "demo shop name")
        self.assertEqual(row[11], "486527013")
        self.assertEqual(row[12], "https://shop486527013.taobao.com/")
        self.assertEqual(row[15], "广东江门")

    def test_export_appends_every_raw_json_leaf_field_after_priority_columns(self):
        export_item_details_to_xlsx(
            db_path=self.db_path,
            output_path=self.xlsx_path,
            num_iids=["599262347474"],
        )

        wb = load_workbook(self.xlsx_path)
        ws = wb["商品信息"]
        headers = [cell.value for cell in ws[1]]
        row = [cell.value for cell in ws[2]]

        self.assertIn("raw_json.api_type", headers)
        self.assertIn("raw_json.server_time", headers)
        self.assertIn("raw_json.item.seller_info.shop_name", headers)
        self.assertIn("raw_json.item.props[0].name", headers)
        self.assertIn("raw_json.item.skus.sku[0].sku_id", headers)
        self.assertEqual(row[headers.index("raw_json.api_type")], "taobao")
        self.assertEqual(row[headers.index("raw_json.item.props[0].name")], "品牌")


if __name__ == "__main__":
    unittest.main()
