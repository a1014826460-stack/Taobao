import json
import os
import sqlite3
import tempfile
import unittest

from openpyxl import load_workbook

from src.tools.export_jd_item_raw_json_to_xlsx import (
    PRIORITY_COLUMNS,
    export_jd_item_details_to_xlsx,
    parse_num_iids,
)


class ExportJDItemRawJsonToXlsxTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.db_path = os.path.join(self.tmp.name, "jd.sqlite3")
        self.xlsx_path = os.path.join(self.tmp.name, "jd.xlsx")
        conn = sqlite3.connect(self.db_path)
        conn.execute(
            """
            CREATE TABLE jd_item_details (
                num_iid TEXT PRIMARY KEY,
                title TEXT,
                price TEXT,
                orginal_price TEXT,
                nick TEXT,
                detail_url TEXT,
                pic_url TEXT,
                brand TEXT,
                cid TEXT,
                shop_id TEXT,
                sales TEXT,
                raw_json TEXT NOT NULL,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        raw_json = {
            "api_type": "jd",
            "item": {
                "num_iid": "10095817869841",
                "title": "京东测试商品",
                "price": "88.00",
                "orginal_price": "99.00",
                "nick": "京东测试店",
                "detail_url": "https://item.jd.com/10095817869841.html",
                "pic_url": "https://img.example.test/jd.jpg",
                "brand": "测试品牌",
                "cid": "123",
                "shop_id": "shop-1",
                "sales": 8,
                "skus": {
                    "sku": [
                        {
                            "sku_id": "sku-1",
                            "price": "88.00",
                            "orginal_price": "99.00",
                            "properties_name": "颜色:红色",
                        }
                    ]
                },
            },
        }
        conn.execute(
            """
            INSERT INTO jd_item_details (
                num_iid, title, price, orginal_price, nick, detail_url,
                pic_url, brand, cid, shop_id, sales, raw_json,
                created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "10095817869841",
                "京东测试商品",
                "88.00",
                "99.00",
                "京东测试店",
                "https://item.jd.com/10095817869841.html",
                "https://img.example.test/jd.jpg",
                "测试品牌",
                "123",
                "shop-1",
                "8",
                json.dumps(raw_json, ensure_ascii=False),
                "2026-07-06T01:00:00+00:00",
                "2026-07-06T02:00:00+00:00",
            ),
        )
        conn.commit()
        conn.close()

    def tearDown(self):
        self.tmp.cleanup()

    def test_parse_num_iids_deduplicates(self):
        self.assertEqual(parse_num_iids(["1,2", "2\n3"]), ["1", "2", "3"])

    def test_export_uses_chinese_headers_and_raw_json_fields(self):
        written = export_jd_item_details_to_xlsx(
            db_path=self.db_path,
            output_path=self.xlsx_path,
            num_iids=["10095817869841"],
        )

        self.assertEqual(written, 1)
        wb = load_workbook(self.xlsx_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        row = [cell.value for cell in ws[2]]

        self.assertEqual(headers[: len(PRIORITY_COLUMNS)], PRIORITY_COLUMNS)
        self.assertEqual(row[0], "京东")
        self.assertEqual(row[1], "10095817869841")
        self.assertEqual(row[2], "京东测试商品")
        self.assertEqual(row[3], "sku-1")
        self.assertEqual(row[4], "颜色:红色")
        self.assertEqual(row[5], "99.00")
        self.assertEqual(row[6], "88.00")
        self.assertIn("raw_json.item.title", headers)
        self.assertIn("raw_json.api_type", headers)
        self.assertEqual(row[headers.index("raw_json.item.title")], "京东测试商品")


if __name__ == "__main__":
    unittest.main()
