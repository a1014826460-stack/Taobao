import json
import sqlite3
import tempfile
import unittest
from pathlib import Path
from unittest import mock

import openpyxl

from src import iqoo_sku_price_backfill as backfill


class IqooSkuPriceBackfillTests(unittest.TestCase):
    def test_phone_model_from_title_recognizes_z_series_x_and_i_variants(self):
        self.assertEqual(
            backfill.phone_model_from_title("vivo iQOO Z11x 手机"),
            "iQOO Z11x",
        )
        self.assertEqual(
            backfill.phone_model_from_title("vivo iQOO Z11i 手机"),
            "iQOO Z11i",
        )
        self.assertEqual(
            backfill.phone_model_from_title("vivo iQOO Z10x 手机"),
            "iQOO Z10x",
        )

    def test_match_sku_requires_exact_model_configuration_and_color(self):
        detail = {
            "item_id": "1001",
            "title": "iQOO 15 phone",
            "sku_base": {
                "props": [
                    {
                        "pid": "p",
                        "name": "版本",
                        "values": [{"vid": "v1", "name": "12+256G"}],
                    },
                    {
                        "pid": "c",
                        "name": "颜色分类",
                        "values": [{"vid": "c1", "name": "传奇版"}],
                    },
                ],
                "skus": [{"propPath": "p:v1;c:c1", "skuId": "s1"}],
            },
            "sku_core": {"sku2info": {"s1": {"price": {"priceText": "2199"}}}},
        }
        match = backfill.match_sku(detail, "iQOO 15", "12+256G", "传奇版")
        self.assertEqual(match.price, "2199")
        self.assertIsNone(backfill.match_sku(detail, "iQOO 15", "16+512G", "传奇版"))

    def test_match_sku_normalizes_gb_and_tb_capacity_units(self):
        detail = {
            "item_id": "1001",
            "title": "iQOO 15 phone",
            "sku_base": {
                "props": [
                    {"pid": "p", "values": [{"vid": "v1", "name": "16GB+1TB"}]},
                    {"pid": "c", "values": [{"vid": "c1", "name": "传奇版"}]},
                ],
                "skus": [{"propPath": "p:v1;c:c1", "skuId": "s1"}],
            },
            "sku_core": {"sku2info": {"s1": {"price": {"priceText": "2199"}}}},
        }
        self.assertEqual(
            backfill.match_sku(detail, "iQOO 15", "16+1T", "传奇版").price,
            "2199",
        )

    def test_match_sku_uses_explicit_page_installment_term(self):
        detail = {
            "item_id": "1001",
            "title": "iQOO 15 phone",
            "installment_periods": 3,
            "sku_base": {
                "props": [
                    {"pid": "p", "values": [{"vid": "v1", "name": "12+256G"}]},
                    {"pid": "c", "values": [{"vid": "c1", "name": "传奇版"}]},
                ],
                "skus": [{"propPath": "p:v1;c:c1", "skuId": "s1"}],
            },
            "sku_core": {"sku2info": {"s1": {"price": {"priceText": "2199"}}}},
        }
        self.assertEqual(
            backfill.match_sku(detail, "iQOO 15", "12+256G", "传奇版").installment_periods,
            3,
        )
        self.assertEqual(backfill.extract_page_installment_periods({"text": "3期免息 约¥850/期起"}), 3)

    def test_build_phone_export_rows_emits_price_and_installment_for_each_sku(self):
        detail = {
            "item_id": "1001",
            "title": "iQOO Z11 Turbo phone",
            "installment_periods": 3,
            "sku_base": {
                "props": [
                    {"pid": "p", "name": "机身颜色", "values": [{"vid": "c1", "name": "极夜黑"}]},
                    {"pid": "m", "name": "存储容量", "values": [{"vid": "v1", "name": "12GB+256GB"}]},
                ],
                "skus": [{"propPath": "p:c1;m:v1", "skuId": "s1"}],
            },
            "sku_core": {"sku2info": {"s1": {"price": {"priceText": "2999"}}}},
        }
        rows = backfill.build_phone_export_rows([detail], "https://iqoo.tmall.com/")
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["机型(*)"], "iQOO Z11 Turbo")
        self.assertEqual(rows[0]["配置(*)"], "12+256G")
        self.assertEqual(rows[0]["颜色(*)"], "极夜黑")
        self.assertEqual(rows[0]["价格"], "2999")
        self.assertEqual(rows[0]["免息分期"], "¥999.67 × 3期")

    def test_build_phone_export_rows_falls_back_to_coupon_price_and_page_installment_text(self):
        detail = {
            "item_id": "1001",
            "title": "iQOO 15 phone",
            "coupon_price": "2199",
            "installment_text": "6期免息 约¥367/期起",
            "sku_base": {
                "props": [
                    {"pid": "p", "name": "机身颜色", "values": [{"vid": "c1", "name": "传奇版"}]},
                    {"pid": "m", "name": "存储容量", "values": [{"vid": "v1", "name": "12GB+256GB"}]},
                ],
                "skus": [{"propPath": "p:c1;m:v1", "skuId": "s1"}],
            },
            "sku_core": {"sku2info": {"s1": {}}},
        }
        rows = backfill.build_phone_export_rows([detail], "https://iqoo.tmall.com/")
        self.assertEqual(rows[0]["价格"], "2199")
        self.assertEqual(rows[0]["免息分期"], "¥367 × 6期")

    def test_build_phone_export_rows_uses_sku_coupon_price_and_per_sku_installment(self):
        detail = {
            "item_id": "1001",
            "title": "iQOO 15 phone",
            "coupon_price": "2199",
            "installment_periods": 6,
            "installment_text": "6期免息 约¥367/期起",
            "sku_base": {
                "props": [
                    {"pid": "p", "name": "版本", "values": [{"vid": "v1", "name": "12+256G"}]},
                    {"pid": "c", "name": "机身颜色", "values": [{"vid": "c1", "name": "传奇版"}]},
                ],
                "skus": [{"propPath": "p:v1;c:c1", "skuId": "s1"}],
            },
            "sku_core": {
                "sku2info": {
                    "s1": {
                        "price": {"priceText": "2599"},
                        "subPrice": {"priceText": "2399"},
                    }
                }
            },
        }

        rows = backfill.build_phone_export_rows([detail], "https://iqoo.tmall.com/")

        self.assertEqual(rows[0]["价格"], "2399")
        self.assertEqual(rows[0]["免息分期"], "¥399.83 × 6期")

    def test_build_phone_export_rows_uses_sku_price_when_coupon_price_is_only_starting_price(self):
        detail = {
            "item_id": "1001",
            "title": "iQOO 15 phone",
            "coupon_price": "2199",
            "sku_base": {
                "props": [
                    {"pid": "p", "name": "版本", "values": [{"vid": "v1", "name": "16+512G"}]},
                    {"pid": "c", "name": "机身颜色", "values": [{"vid": "c1", "name": "传奇版"}]},
                ],
                "skus": [{"propPath": "p:v1;c:c1", "skuId": "s1"}],
            },
            "sku_core": {"sku2info": {"s1": {"price": {"priceText": "3899"}}}},
        }

        rows = backfill.build_phone_export_rows([detail], "https://iqoo.tmall.com/")

        self.assertEqual(rows[0]["价格"], "3899")

    def test_build_phone_export_rows_uses_real_sku_discounted_prices_when_captured(self):
        details = backfill.load_detail_records("data/taobao_items.sqlite3")

        rows = backfill.build_phone_export_rows(details, backfill.DEFAULT_SHOP_URL)

        turbo_prices = {
            (row["配置(*)"], row["颜色(*)"]): row["价格"]
            for row in rows
            if row["机型(*)"] == "iQOO Z11 Turbo"
        }
        self.assertEqual(turbo_prices[("12+256G", "天光白")], "2549.15")
        self.assertEqual(turbo_prices[("16+256G", "天光白")], "2804.15")
        self.assertEqual(turbo_prices[("16+512G", "天光白")], "3399")
        self.assertEqual(turbo_prices[("16+1T", "天光白")], "4599")

    def test_build_phone_export_rows_keeps_first_row_for_duplicate_identity(self):
        first = {
            "item_id": "1001",
            "title": "iQOO 15 phone",
            "coupon_price": "2199",
            "sku_base": {
                "props": [
                    {"pid": "p", "name": "版本", "values": [{"vid": "v1", "name": "12+256G"}]},
                    {"pid": "c", "name": "机身颜色", "values": [{"vid": "c1", "name": "传奇版"}]},
                ],
                "skus": [{"propPath": "p:v1;c:c1", "skuId": "s1"}],
            },
            "sku_core": {"sku2info": {"s1": {}}},
        }
        duplicate = {**first, "item_id": "1002", "coupon_price": "2099"}

        rows = backfill.build_phone_export_rows(
            [first, duplicate], "https://iqoo.tmall.com/"
        )

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["价格"], "2199")

    def test_write_phone_export_writes_requested_headers(self):
        rows = [{header: header for header in backfill.PHONE_EXPORT_HEADERS}]
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "phones.xlsx"
            written = backfill.write_phone_export(output, rows)
            workbook = openpyxl.load_workbook(output, read_only=True, data_only=True)
            try:
                sheet = workbook.active
                self.assertEqual(written, 1)
                self.assertEqual([cell.value for cell in sheet[1]], backfill.PHONE_EXPORT_HEADERS)
                self.assertEqual([cell.value for cell in sheet[2]], backfill.PHONE_EXPORT_HEADERS)
            finally:
                workbook.close()

    def test_build_item_get_phone_export_rows_uses_sku_price_and_null_fallbacks(self):
        records = [
            {
                "num_iid": "1001",
                "title": "vivo iQOO Z11 Turbo 手机",
                "nick": "iQOO手机官方旗舰店",
                "detail_url": "https://item.taobao.com/item.htm?id=1001",
                "brand": "vivo",
                "skus": {
                    "sku": [
                        {
                            "price": 2699,
                            "orginal_price": 2799,
                            "properties_name": "机身颜色:天光白;储存容量:12GB+256GB",
                        },
                        {
                            "price": None,
                            "properties_name": "机身颜色:极夜黑;储存容量:16GB+512GB",
                        },
                    ]
                },
            }
        ]

        rows = backfill.build_item_get_phone_export_rows(records)

        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["店铺"], "iQOO手机官方旗舰店")
        self.assertEqual(rows[0]["网页版链接"], "https://item.taobao.com/item.htm?id=1001")
        self.assertEqual(rows[0]["品牌(*)"], "vivo")
        self.assertEqual(rows[0]["机型(*)"], "iQOO Z11 Turbo")
        self.assertEqual(rows[0]["配置(*)"], "12+256G")
        self.assertEqual(rows[0]["颜色(*)"], "天光白")
        self.assertEqual(rows[0]["原价"], "2799")
        self.assertEqual(rows[0]["优惠价(券后价)"], "2699")
        self.assertEqual(rows[0]["免息分期"], "NULL")
        self.assertEqual(rows[1]["原价"], "NULL")
        self.assertEqual(rows[1]["优惠价(券后价)"], "NULL")

    def test_build_item_get_phone_export_rows_deduplicates_identical_phone_combinations(self):
        record = {
            "num_iid": "1001",
            "title": "vivo iQOO Z11 Turbo 手机",
            "nick": "iQOO手机官方旗舰店",
            "detail_url": "https://item.taobao.com/item.htm?id=1001",
            "brand": "vivo",
            "skus": {
                "sku": [
                    {
                        "price": 2699,
                        "properties_name": "机身颜色:天光白;储存容量:12GB+256GB",
                    },
                    {
                        "price": 2699,
                        "properties_name": "机身颜色:天光白;储存容量:12GB+256GB",
                    },
                ]
            },
        }

        rows = backfill.build_item_get_phone_export_rows([record, record])

        self.assertEqual(len(rows), 1)

    def test_format_installment_and_load_shop_candidates(self):
        self.assertEqual(backfill.format_installment("2159", 12), "¥179.92 × 12期")
        self.assertEqual(backfill.format_installment("2159", None), "")
        with tempfile.TemporaryDirectory() as directory:
            db_path = Path(directory) / "shop.sqlite3"
            db = sqlite3.connect(db_path)
            db.executescript(
                """
                CREATE TABLE tmall_shop_items (shop_url TEXT, item_id TEXT, title TEXT);
                INSERT INTO tmall_shop_items VALUES
                    ('https://iqoo.tmall.com/search.htm', '1001', 'iQOO 15 phone'),
                    ('https://iqoo.tmall.com/search.htm', '1002', 'iQOO Pad');
                """
            )
            db.commit()
            db.close()
            self.assertEqual(
                backfill.load_shop_candidates(
                    db_path, "https://iqoo.tmall.com/search.htm", "iQOO 15"
                ),
                ["1001"],
            )

    def test_load_shop_candidates_ignores_query_parameters_in_shop_url(self):
        with tempfile.TemporaryDirectory() as directory:
            db_path = Path(directory) / "shop.sqlite3"
            db = sqlite3.connect(db_path)
            db.executescript(
                """
                CREATE TABLE tmall_shop_items (shop_url TEXT, item_id TEXT, title TEXT);
                INSERT INTO tmall_shop_items VALUES
                    ('https://iqoo.tmall.com/search.htm', '1001', 'iQOO 15 phone');
                """
            )
            db.commit()
            db.close()
            self.assertEqual(
                backfill.load_shop_candidates(
                    db_path,
                    "https://iqoo.tmall.com/search.htm?orderType=defaultSort",
                    "iQOO 15",
                ),
                ["1001"],
            )

    def test_load_shop_candidates_matches_shop_url_despite_query_order(self):
        with tempfile.TemporaryDirectory() as directory:
            db_path = Path(directory) / "shop.sqlite3"
            db = sqlite3.connect(db_path)
            db.executescript(
                """
                CREATE TABLE tmall_shop_items (shop_url TEXT, item_id TEXT, title TEXT);
                INSERT INTO tmall_shop_items VALUES
                    ('https://iqoo.tmall.com/search.htm?orderType=defaultSort&viewType=grid', '1001', 'iQOO 15 phone');
                """
            )
            db.commit()
            db.close()
            self.assertEqual(
                backfill.load_shop_candidates(
                    db_path,
                    "https://iqoo.tmall.com/search.htm?viewType=grid&orderType=defaultSort",
                    "iQOO 15",
                ),
                ["1001"],
            )

    def test_load_shop_candidates_falls_back_to_raw_shop_page_titles(self):
        with tempfile.TemporaryDirectory() as directory:
            db_path = Path(directory) / "shop.sqlite3"
            db = sqlite3.connect(db_path)
            db.executescript(
                """
                CREATE TABLE tmall_shop_items (shop_url TEXT, item_id TEXT, title TEXT);
                CREATE TABLE tmall_shop_pages (shop_url TEXT, page_number INTEGER, raw_json TEXT);
                INSERT INTO tmall_shop_items VALUES
                    ('https://iqoo.tmall.com/search.htm', '1001', '');
                """
            )
            raw_html = '''<dl class="item" data-id="1001"><dd class="detail"><a class="item-name">iQOO 15 phone</a></dd></dl>'''
            db.execute(
                "INSERT INTO tmall_shop_pages VALUES (?, ?, ?)",
                ("https://iqoo.tmall.com/search.htm", 1, json.dumps(raw_html)),
            )
            db.commit()
            db.close()
            self.assertEqual(
                backfill.load_shop_candidates(
                    db_path, "https://iqoo.tmall.com/search.htm", "iQOO 15"
                ),
                ["1001"],
            )

    def test_write_backfill_workbooks_writes_exact_prices_and_report(self):
        detail_with_installment = {
            "item_id": "1001",
            "title": "iQOO 15 phone",
            "sku_base": {
                "props": [
                    {"pid": "p", "values": [{"vid": "v1", "name": "12+256G"}]},
                    {"pid": "c", "values": [{"vid": "c1", "name": "传奇版"}]},
                ],
                "skus": [{"propPath": "p:v1;c:c1", "skuId": "s1"}],
            },
            "sku_core": {
                "sku2info": {"s1": {"price": {"priceText": "2199"}, "installmentPeriods": 12}}
            },
        }
        detail_without_installment = {
            "item_id": "1002",
            "title": "iQOO Neo11 phone",
            "sku_base": {
                "props": [
                    {"pid": "p", "values": [{"vid": "v1", "name": "16+512G"}]},
                    {"pid": "c", "values": [{"vid": "c1", "name": "疾影黑"}]},
                ],
                "skus": [{"propPath": "p:v1;c:c1", "skuId": "s1"}],
            },
            "sku_core": {"sku2info": {"s1": {"price": {"priceText": "999"}}}},
        }
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "source.xlsx"
            output = Path(directory) / "filled.xlsx"
            report = Path(directory) / "report.xlsx"
            book = openpyxl.Workbook()
            sheet = book.active
            sheet.append(["店铺", "网页版链接", "品牌(*)", "机型(*)", "配置(*)", "颜色(*)", "价格", "免息分期"])
            sheet.append(["shop", "url", "iQOO", "iQOO 15", "12+256G", "传奇版", None, None])
            sheet.append(["shop", "url", "iQOO", "iQOO Neo11", "16+512G", "疾影黑", None, None])
            sheet.append(["shop", "url", "iQOO", "iQOO Z11", "12+256G", "黑", None, None])
            book.save(source)

            result = backfill.write_backfill_workbooks(
                source, output, report, [detail_with_installment, detail_without_installment]
            )
            filled = openpyxl.load_workbook(output).active
            report_sheet = openpyxl.load_workbook(report).active
            self.assertEqual(filled["G2"].value, "2199")
            self.assertEqual(filled["H2"].value, "¥183.25 × 12期")
            self.assertEqual(filled["G3"].value, "999")
            self.assertIsNone(filled["H3"].value)
            self.assertIsNone(filled["G4"].value)
            self.assertEqual(result["filled_prices"], 2)
            self.assertEqual(report_sheet.max_row, 3)

    def test_supplement_details_pauses_after_first_crawler_failure(self):
        with mock.patch.object(backfill, "load_shop_candidates", return_value=["1001", "1002"]):
            crawler = mock.Mock(return_value=1)
            result = backfill.supplement_details(
                models=["iQOO 15", "iQOO Neo11"],
                shop_db="shop.sqlite3",
                detail_db="details.sqlite3",
                shop_url="https://iqoo.tmall.com/search.htm",
                crawler=crawler,
            )
        self.assertTrue(result.paused)
        self.assertEqual(result.attempted_ids, ["1001"])
        self.assertEqual(crawler.call_count, 1)


if __name__ == "__main__":
    unittest.main()
