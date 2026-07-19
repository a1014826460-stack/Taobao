import json
import sqlite3
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from src.jd_ware_business_store import (
    CHINESE_COLUMNS,
    SQLiteJDWareBusinessStore,
    build_export_row,
    export_to_xlsx,
)


class JDWareBusinessStoreTest(unittest.TestCase):
    def sample_response(self):
        return {
            "skuHeadVO": {"skuName": "测试商品标题"},
            "price": {
                "p": "63.90",
                "op": "69.90",
                "finalPrice": {"price": "60.90", "priceContent": "到手价"},
            },
            "itemShopInfo": {
                "shopName": "测试店铺",
                "shopId": "12345",
                "venderId": "67890",
            },
            "stockVO": {"stockStateDesc": "现货", "areaName": "广东广州"},
            "wareInfo": {"wareInfoMap": {"sku_status": "1"}},
            "promotion": {"activity": ["满减"]},
            "isLogin": True,
        }

    def test_build_export_row_uses_chinese_columns(self):
        row = build_export_row("10207466352379", self.sample_response(), "2026-07-08T00:00:00Z")
        self.assertEqual(row["平台名称"], "京东")
        self.assertEqual(row["商品ID"], "10207466352379")
        self.assertEqual(row["spu名称"], "测试商品标题")
        self.assertEqual(row["正常价格（标价）"], "63.90")
        self.assertEqual(row["到手价"], "60.90")
        self.assertEqual(row["店铺名称"], "测试店铺")
        self.assertEqual(row["发货地区"], "广东广州")

    def test_store_save_and_export_xlsx(self):
        with tempfile.TemporaryDirectory() as tmp:
            db_path = Path(tmp) / "items.sqlite3"
            xlsx_path = Path(tmp) / "items.xlsx"
            store = SQLiteJDWareBusinessStore(db_path)
            store.save_success("10207466352379", self.sample_response(), "https://api.m.jd.com/?functionId=x", 200)
            store.close()

            conn = sqlite3.connect(db_path)
            try:
                row = conn.execute("SELECT num_iid, title, price, raw_json FROM jd_ware_business_details").fetchone()
                self.assertEqual(row[0], "10207466352379")
                self.assertEqual(row[1], "测试商品标题")
                self.assertEqual(row[2], "60.90")
                self.assertIn("测试商品标题", row[3])
            finally:
                conn.close()

            written = export_to_xlsx(db_path, xlsx_path, ["10207466352379"])
            self.assertEqual(written, 1)
            wb = load_workbook(xlsx_path)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            self.assertEqual(headers[: len(CHINESE_COLUMNS)], CHINESE_COLUMNS)
            values = dict(zip(headers, [cell.value for cell in ws[2]]))
            self.assertEqual(values["商品ID"], "10207466352379")
            self.assertEqual(values["店铺名称"], "测试店铺")


if __name__ == "__main__":
    unittest.main()
